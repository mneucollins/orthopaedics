""" Script en python para filtrar resultados de un archivo de excel
		y volcarlos en otro archivo.
		este script requiere 3 parametros de entrada:
			-Archivo de origen
			-Cadena de busqueda
			-Archivo de salida
	Version 0.1 a partir de las modificaciones de kiocone
	Version 0.2 abre la plantilla 'macros.xlsm' paravaciar los datos
		del resultado y gardarlo como un archivo de macros nuevo"""
from openpyxl import load_workbook, Workbook
from datetime import datetime
import sys
import warnings
import operator
import copy
import math

class doc:
    def __init__(self, name = None, apt= None, time=None, wr=None, ex=None):
    	self.name = name
    	self.apt = apt
    	self.time = time
    	self.wr = wr
    	self.ex = ex

warnings.simplefilter("ignore")
#Carga plantilla con macros
try:
	wbnew = load_workbook("/Users/ezabaw/Documents/orthopaedics/reports/macros.xlsm", read_only=False, keep_vba=True)
	print "Template file loaded"
except:
	print "Error macros template"
	sys.exit()
#si el nombre de origen tiene mas de 5 caracteres incluida la extension de archivo
if len(sys.argv) == 4:
	#si el nombre del archivo tiene al menos 6 caracteres pasa(incluyendo la extension)
	if len(sys.argv[1]) > 5:
		try:
			#Carga archivo indicado en el argumento 1
			print "loading..."
			wb = load_workbook(sys.argv[1])
			print "load complete"
		except:
			print "The file", str(sys.argv[1]),"does not exist or is misspelled"
			sys.exit()
		sheet_ranges = wb['PatientStats']
		DoctorName = str(sys.argv[2])
		if len(sys.argv[3]) < 6:
			print "Output file too short"
			sys.exit()
		else:
			SummaryFile = str(sys.argv[3])
	else:
		print "Invalid file name"
		sys.exit()
else:
    print "this script requires 3 arguments"
    print "1- report file name"
    print "2- Doctor name, single or double quoted"
    print "3- Output file name with 'xlsm' extension"
    sys.exit()

#rellena en la variable info la matriz de sheet_ranges
j = 0
info = []
for row in sheet_ranges.rows:
	j =j + 1
	if (j > 1):
		i = 1
		do = True
		for cell in row:
			try:
				if (i == 3):
					name = str(cell.value)
					if (name != DoctorName):
						do = False
				elif (i == 6):
					apt = str(cell.value)
					apt = datetime.strptime(apt.split('.')[0], '%H:%M:%S')
				elif (i == 16):
					time = str(cell.value)
				elif (i == 17):
					wr = str(cell.value)
					if (int(wr) >= 400):
						do = False
				elif (i == 18):
					ex = str(cell.value)
					if (int(ex) <= 10):
						do = False
			except:
				do = False
			i = i + 1
		if do:
			t = doc(name, apt, time, wr, ex)
			info.append(t)

#si la longitud de la matriz info es 0 termina
if len(info) == 0:
    print "No valid information found for doctor. Please check if valid Callback vs. Appt Time present"
else:
	info.sort(key=operator.attrgetter('apt'))
	first = info[0].apt
	last = info[len(info) - 1].apt

	firstList = []
	lastList = []
	for i in info:
		c = i.apt
		if (c.hour == first.hour and c.minute >= first.minute):
			firstList.append(i)
		elif (c.hour == first.hour + 1 and c.minute < first.minute):
			firstList.append(i)

		if (c.hour == last.hour and c.minute <= last.minute):
			lastList.append(i)
		elif (c.hour == last.hour - 1 and c.minute == last.minute and c.second > last.second or c.hour == last.hour - 1 and c.minute > last.minute):
			lastList.append(i)

	sheet_ranges1 = wbnew.active

	t = 0

	for a in lastList:
		t = int(a.time) + t
	avgM = float(t)/len(lastList)
	st = 0
	for a in lastList:
		if len(lastList) > 1:
			st = math.pow((math.pow((int(a.time) - avgM), 2)) / (len(lastList) - 1), 0.5)
		else:
			st = 0

	lastHourB = (avgM)
	lastHourC = (st)
	lastHourD = len(lastList)

	t = 0
	for a in firstList:
		t = int(a.time) + t
	avgM = float(t)/len(firstList)
	st = 0
	for a in firstList:
		if len(firstList) > 1:
			st = math.pow((math.pow((int(a.time) - avgM), 2)) / (len(firstList) - 1), 0.5)
		else:
			st = 0

	firstHourB = (avgM)
	firstHourC = (st)
	firstHourD = len(firstList)


	for i in info:
		if i.apt.minute >= 30:
			i.apt = i.apt.replace(minute = 30, second = 0)
		else:
			i.apt = i.apt.replace(minute = 0, second = 0)

	dc = {}

	for i in info:
		if i.apt.hour < 10:
			h = "0" + str(i.apt.hour)
		else:
			h = str(i.apt.hour)
		if (i.apt.minute == 0):
			myDate = h + ":00:00 - " + h + ":29:59"
		else:
			myDate = h + ":30:00 - " + h + ":59:59"
		if myDate in dc:
			dc[myDate].append(i)
		else:
			dc[myDate] = [i]

	morningValue = []
	afternoonValue = []
	index = 3
	for k in sorted(dc):
		val = dc[k]
		total = 0
		for v in val:
			total = total + int(v.time)
		avg = float(total)/(len(val))
		stotal = 0
		count = len(val)
		for v in val:
			stotal = stotal + math.pow((int(v.time) - avg), 2)
		if (count > 1):
			stotal = math.pow((stotal / (count - 1)), 0.5)
		else:
			stotal = 0

		if val[0].apt.hour < 12:
			morningValue = morningValue + val
		else:
			afternoonValue = afternoonValue + val

		sheet_ranges1['A' + str(index)] = k
		sheet_ranges1['B' + str(index)] = avg
		sheet_ranges1['C' + str(index)] = stotal
		sheet_ranges1['D' + str(index)] = count


		index = index + 1

		# print "Interval: " + str(k)
		# print "Average: " + str(avg)
		# print "Count: " + str(count)
		# print "Standard Dev: " + str(stotal)
		# print " "

	totalValue = morningValue + afternoonValue

	sheet_ranges1['A1'] = str(DoctorName)

	sheet_ranges1.title = "Summary"
	sheet_ranges1['A2'] = "Appt Time"
	sheet_ranges1['B2'] = "Average of Callback vs. Appt Time"
	sheet_ranges1['C2'] = "StdDev of Callback vs. Appt Time"
	sheet_ranges1['D2'] = "Count of Callback vs. Appt Time"

	sheet_ranges1['A' + str(index)] = "Morning"
	t = 0
	for a in morningValue:
		t = int(a.time) + t
	avgM = float(t)/len(morningValue)
	st = 0
	for a in morningValue:
		if len(morningValue) > 1:
			st = math.pow((math.pow((int(a.time) - avgM), 2)) / (len(morningValue) - 1), 0.5)
		else:
			st = 0

	sheet_ranges1['B' + str(index)] = (avgM)
	sheet_ranges1['C' + str(index)] = (st)
	sheet_ranges1['D' + str(index)] = len(morningValue)
	index = index + 1


	sheet_ranges1['A' + str(index)] = "Afternoon"
	t = 0
	for a in afternoonValue:
		t = int(a.time) + t
	avgM = float(t)/len(afternoonValue)
	st = 0
	for a in afternoonValue:
		if len(afternoonValue) > 1:
			st = math.pow((math.pow((int(a.time) - avgM), 2)) / (len(afternoonValue) - 1), 0.5)
		else:
			st = 0

	sheet_ranges1['B' + str(index)] = (avgM)
	sheet_ranges1['C' + str(index)] = (st)
	sheet_ranges1['D' + str(index)] = len(afternoonValue)

	index = index + 1

	sheet_ranges1['A' + str(index)] = "First Hour"
	sheet_ranges1['B' + str(index)] = firstHourB
	sheet_ranges1['C'  + str(index)] = firstHourC
	sheet_ranges1['D' + str(index)] = firstHourD
	index = index + 1

	sheet_ranges1['A' + str(index)] = "Last Hour"
	sheet_ranges1['B' + str(index)] = lastHourB
	sheet_ranges1['C' + str(index)] = lastHourC
	sheet_ranges1['D' + str(index)] = lastHourD

	index = index + 1
	sheet_ranges1['A' + str(index)] = "Grand Total"
	t = 0
	for a in totalValue:
		t = int(a.time) + t
	avgM = float(t)/len(totalValue)
	st = 0
	for a in totalValue:
		if len(totalValue) > 1:
			st = math.pow((math.pow((int(a.time) - avgM), 2)) / (len(totalValue) - 1), 0.5)
		else:
			st = 0

	sheet_ranges1['B' + str(index)] = (avgM)
	sheet_ranges1['C' + str(index)] = (st)
	sheet_ranges1['D' + str(index)] = len(totalValue)

	wbnew.save(SummaryFile)

	print "Saved to", SummaryFile
